#!/usr/bin/env python3
"""Process New Referrals Batch 2: Excel + OneNote -> verified SQL.

Phases:
  A. Load & clean Excel spreadsheet
  B. Resolve folder names to disk folders
  C. Extract OneNote demographics
  D. Cross-reference & discrepancy report
  E. Generate SQL (mhs001_mpi, mhs002_gp, mhs101_referral, mhs102_service_type_referred_to)

Usage:
    uv run process_batch2.py "/Users/georgi/Downloads/New Referrals Batch 2"
"""

import csv
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from extract_patient_records import (
    extract_demographics,
    extract_utf16le_strings,
    find_author_names,
)
from generate_sql import parse_address, split_name, sql_val

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ICB name -> code mapping (from ref_icb seed data)
ICB_LOOKUP = {
    "NHS North East and North Cumbria ICB": "QHM",
    "NHS Norfolk and Waveney ICB": "QMM",
    "NHS Northamptonshire ICB": "QPM",
    "NHS Lincolnshire ICB": "QJM",
    "NHS Lancashire and South Cumbria ICB": "QE1",
    "NHS Cornwall and the Isles of Scilly ICB": "QT6",
    "NHS Cornwall ICB": "QT6",
    "NHS Gloucestershire ICB": "QR1",
    "NHS Somerset ICB": "QSL",
    "NHS Greater Manchester ICB": "QOP",
    "NHS Dorset ICB": "QVV",
    "NHS Derby and Derbyshire ICB": "QJ2",
    "NHS North West London ICB": "QRV",
    "NHS Cheshire and Merseyside ICB": "QYG",
    "NHS Birmingham and Solihull ICB": "QHL",
    "NHS Black Country ICB": "QUA",
    "NHS Birmingham and Black Country ICB": "QUA",
    "NHS Shropshire, Telford and Wrekin ICB": "QOC",
    "NHS West Yorkshire ICB": "QWO",
    "NHS Surrey Heartlands ICB": "QXU",
    "NHS Devon ICB": "QJK",
    "NHS Frimley ICB": "QNQ",
    "NHS Frimley": "QNQ",
    "NHS Staffordshire and Stoke-on-Trent ICB": "QNC",
    "NHS Bath and North East Somerset, Swindon and Wiltshire ICB": "QOX",
    "NHS Bristol, North Somerset and South Gloucestershire ICB": "QUY",
    "NHS Coventry and Warwickshire ICB": "QWU",
    "NHS Sussex ICB": "QNX",
    "NHS Buckinghamshire, Oxfordshire and Berkshire West ICB": "QU9",
    "NHS Buckinghamshire Oxfordshire and Berkshire West ICB": "QU9",
    "NHS Hampshire and Isle of Wight ICB": "QRL",
    "NHS Kent and Medway ICB": "QKS",
    "NHS Hertfordshire and West Essex ICB": "QM7",
    "NHS Leicester, Leicestershire and Rutland ICB": "QK1",
    "NHS Cambridgeshire and Peterborough ICB": "QUE",
    "NHS Bedfordshire, Luton and Milton Keynes ICB": "QHG",
    "NHS Mid and South Essex ICB": "QH8",
    "NHS Suffolk and North East Essex ICB": "QJG",
    "NHS South Yorkshire ICB": "QF7",
    "NHS Nottingham and Nottinghamshire ICB": "QT1",
    "NHS North Central London ICB": "QMJ",
    "NHS South West London ICB": "QWE",
    "NHS Humber and North Yorkshire ICB": "QOQ",
    "NHS Herefordshire and Worcestershire ICB": "QGH",
    "NHS North East London ICB": "QMF",
    "NHS South East London ICB": "QKK",
    "NHS Midlands ICB": "QHL",
    "NHS Trafford ICB": "QOP",
    "North West London ICB": "QRV",
}

EXCEL_COLUMNS = [
    "Patient Name",
    "DOB",
    "Patient Email",
    "Patient Phone Number",
    "Address Line 1",
    "Address Line 2",
    "City",
    "Post Code",
    "NHS Number",
    "GP Practice Name",
    "Referrer Organisation",
    "Referrer Name",
    "Referrer Email",
    "GP Practice Email",
    "Referral Date",
    "ICB",
    "Folder Name",
    "Status/Notes",
]


# ── Phase A: Load & clean Excel ─────────────────────────────────────────────


def clean_string(val):
    """Strip whitespace and non-breaking spaces from a string value."""
    if val is None:
        return ""
    s = str(val).replace("\xa0", " ").strip()
    return s


def normalise_dob(val):
    """Convert DOB to DD/MM/YYYY string. Handles datetime objects and strings."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = clean_string(val)
    # Already DD/MM/YYYY?
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return s
    return s


def normalise_nhs(val):
    """Normalise NHS number to 10-digit string without spaces."""
    s = clean_string(val).replace(" ", "")
    if re.match(r"^\d{10}$", s):
        return s
    return s


def load_excel(xlsx_path):
    """Load and clean the referral register Excel file."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean_string(row[0])
        if not name:
            continue

        record = {}
        for i, col in enumerate(EXCEL_COLUMNS):
            record[col] = clean_string(row[i]) if i < len(row) else ""

        # Normalise specific fields
        record["DOB"] = normalise_dob(row[1])
        record["NHS Number"] = normalise_nhs(row[8])
        record["Patient Phone Number"] = clean_string(row[3])
        record["Patient Email"] = clean_string(row[2])

        patients.append(record)

    wb.close()
    logger.info(f"Phase A: Loaded {len(patients)} patients from Excel")
    return patients


# ── Phase B: Resolve folder names ───────────────────────────────────────────


def resolve_folders(patients, files_dir):
    """Map Excel Folder Name values to actual disk folders."""
    if not files_dir.is_dir():
        logger.warning(f"Patient Files directory not found: {files_dir}")
        return

    disk_folders = {f.name: f for f in files_dir.iterdir() if f.is_dir()}

    matched = 0
    prefix_matched = 0
    unmatched = 0

    for p in patients:
        folder_name = p["Folder Name"]
        if not folder_name:
            p["_folder_path"] = None
            unmatched += 1
            continue

        # Exact match
        if folder_name in disk_folders:
            p["_folder_path"] = disk_folders[folder_name]
            matched += 1
            continue

        # Prefix match (for folders missing DOB suffix)
        candidates = [
            name for name in disk_folders if name.startswith(folder_name)
        ]
        if len(candidates) == 1:
            p["_folder_path"] = disk_folders[candidates[0]]
            p["Folder Name"] = candidates[0]  # Update to actual folder name
            prefix_matched += 1
            logger.info(f"  Prefix match: {folder_name} -> {candidates[0]}")
        elif len(candidates) > 1:
            logger.warning(
                f"  Ambiguous prefix match for {folder_name}: {candidates}"
            )
            p["_folder_path"] = None
            unmatched += 1
        else:
            logger.warning(f"  No folder match for: {folder_name}")
            p["_folder_path"] = None
            unmatched += 1

    logger.info(
        f"Phase B: {matched} exact, {prefix_matched} prefix, {unmatched} unmatched"
    )


# ── Phase C: Extract OneNote data ───────────────────────────────────────────


def extract_onenote(patients):
    """Extract demographics from OneNote files for each patient."""
    extracted = 0
    skipped = 0

    for p in patients:
        folder = p.get("_folder_path")
        if folder is None or not folder.is_dir():
            p["_onenote"] = {}
            p["_gp_onenote"] = {}
            skipped += 1
            continue

        # Find Personal Patient Records subfolder (flexible name)
        records_dir = None
        for sub in folder.iterdir():
            if sub.is_dir() and sub.name.startswith("Personal Patient Records"):
                records_dir = sub
                break

        if records_dir is None:
            p["_onenote"] = {}
            p["_gp_onenote"] = {}
            skipped += 1
            continue

        demographics_file = records_dir / "Untitled Section.one"
        treatment_file = records_dir / "Treatment Notes.one"

        # Detect author names
        all_strings = []
        for fp in (demographics_file, treatment_file):
            if fp.exists():
                all_strings.extend(extract_utf16le_strings(fp))
        author_names = find_author_names(all_strings)

        if demographics_file.exists():
            demographics, gp_info = extract_demographics(
                demographics_file, author_names
            )
            p["_onenote"] = demographics
            p["_gp_onenote"] = gp_info
            extracted += 1
        else:
            p["_onenote"] = {}
            p["_gp_onenote"] = {}
            skipped += 1

    logger.info(f"Phase C: {extracted} extracted, {skipped} skipped")


# ── Phase D: Cross-reference & discrepancy report ───────────────────────────


def normalise_for_compare(field, value):
    """Normalise a value for comparison."""
    if not value:
        return ""
    s = str(value).strip().lower().replace("\xa0", " ")
    if field in ("DOB", "Date of Birth"):
        s = s.replace("-", "/")
    if field in ("NHS Number",):
        s = s.replace(" ", "")
    if field in ("Phone", "Telephone Number", "Patient Phone Number"):
        s = re.sub(r"[^\d]", "", s).lstrip("0")
    if field in ("Email", "GP Email"):
        s = re.sub(r"^mailto:", "", s)  # strip mailto: prefix
        s = s.strip().rstrip(",>.;")    # strip trailing punctuation artefacts
    if field == "GP Practice":
        s = s.replace("practise", "practice")  # normalise spelling variant
    return s


COMPARE_FIELDS = [
    ("DOB", "DOB", "Date of Birth"),
    ("NHS Number", "NHS Number", "NHS Number"),
    ("Email", "Patient Email", "Email Address"),
    ("Phone", "Patient Phone Number", "Telephone Number"),
    ("GP Practice", "GP Practice Name", "GP Practice"),
    ("GP Email", "GP Practice Email", "GP Email"),
]


def cross_reference(patients, output_dir):
    """Compare Excel vs OneNote field-by-field and write discrepancy report."""
    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "discrepancy_report.csv"

    rows = []
    total_compared = 0
    total_mismatches = 0

    for p in patients:
        onenote = p.get("_onenote", {})
        gp_onenote = p.get("_gp_onenote", {})
        if not onenote and not gp_onenote:
            continue

        total_compared += 1
        all_onenote = {**onenote, **gp_onenote}

        for label, excel_key, onenote_key in COMPARE_FIELDS:
            excel_val = p.get(excel_key, "")
            onenote_val = all_onenote.get(onenote_key, "")

            if not excel_val and not onenote_val:
                continue

            norm_excel = normalise_for_compare(label, excel_val)
            norm_onenote = normalise_for_compare(label, onenote_val)

            if label == "GP Practice":
                match = norm_excel == norm_onenote or (
                    norm_excel and norm_onenote and (
                        norm_excel in norm_onenote or norm_onenote in norm_excel
                    )
                )
            else:
                match = norm_excel == norm_onenote
            if not match and norm_excel and norm_onenote:
                total_mismatches += 1
                rows.append({
                    "Folder Name": p["Folder Name"],
                    "Patient Name": p["Patient Name"],
                    "Field": label,
                    "Excel Value": excel_val,
                    "OneNote Value": onenote_val,
                    "Match": "No",
                })

    with open(report_path, "w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "Folder Name", "Patient Name", "Field",
                "Excel Value", "OneNote Value", "Match",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    logger.info(
        f"Phase D: {total_compared} patients compared, "
        f"{total_mismatches} mismatches -> {report_path}"
    )


# ── Phase E: Generate SQL ───────────────────────────────────────────────────


def parse_referral_date(date_str):
    """Convert DD/MM/YYYY referral date to YYYY-MM-DD."""
    if not date_str:
        return None
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", date_str.strip())
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None


def parse_dob_to_iso(dob_str):
    """Convert DD/MM/YYYY DOB to YYYY-MM-DD for SQL."""
    return parse_referral_date(dob_str)


def get_discrepancy_folders(report_path):
    """Read discrepancy report and return set of Folder Names with mismatches."""
    folders = set()
    if not report_path.is_file():
        return folders
    with open(report_path, newline="") as f:
        for row in csv.DictReader(f):
            folders.add(row["Folder Name"])
    return folders


def generate_sql(patients, output_dir, suffix=""):
    """Generate verified SQL for all four tables."""
    output_dir.mkdir(parents=True, exist_ok=True)
    sql_path = output_dir / f"seed_patients_batch2{suffix}.sql"

    lines = [
        "-- Batch 2 patient seed data",
        "-- Generated from referral_register_batch2.xlsx + OneNote cross-reference",
        "-- Review before running against the database",
        "",
        "BEGIN;",
        "",
        "-- =============================================================",
        "-- mhs001_mpi  (Master Patient Index)",
        "-- =============================================================",
        "",
    ]

    skipped = []
    ref_num = 0

    # Collect referral info for later sections
    referral_data = []

    for p in patients:
        nhs = p["NHS Number"]
        if not nhs or not re.match(r"^\d{10}$", nhs):
            skipped.append((p["Patient Name"], nhs or "MISSING"))
            continue

        onenote = p.get("_onenote", {})

        first_name, last_name = split_name(p["Patient Name"])
        dob = parse_dob_to_iso(p["DOB"])

        # Excel is primary, OneNote fills gaps
        phone = p.get("Patient Phone Number") or onenote.get("Telephone Number") or None
        email = p.get("Patient Email") or onenote.get("Email Address") or None
        address_1 = p.get("Address Line 1") or None
        address_2 = p.get("Address Line 2") or None
        city = p.get("City") or None
        postcode = p.get("Post Code") or None

        # If no structured address from Excel but OneNote has a full address, parse it
        if not address_1 and onenote.get("Address"):
            addr_line, pc = parse_address(onenote["Address"])
            address_1 = addr_line
            if not postcode:
                postcode = pc

        # Normalise postcode spacing
        if postcode:
            pc = postcode.replace(" ", "").upper()
            if len(pc) >= 5:
                postcode = pc[:-3] + " " + pc[-3:]

        lines.append(
            f"INSERT INTO mhs001_mpi "
            f"(nhs_number, first_name, last_name, date_of_birth, "
            f"phone_number, email, address_line_1, address_line_2, city, postcode)\n"
            f"VALUES ({sql_val(nhs)}, {sql_val(first_name)}, {sql_val(last_name)}, "
            f"{sql_val(dob)}, {sql_val(phone)}, {sql_val(email)}, "
            f"{sql_val(address_1)}, {sql_val(address_2)}, {sql_val(city)}, "
            f"{sql_val(postcode)})\n"
            f"ON CONFLICT (nhs_number) DO UPDATE SET\n"
            f"  first_name = EXCLUDED.first_name,\n"
            f"  last_name = EXCLUDED.last_name,\n"
            f"  date_of_birth = EXCLUDED.date_of_birth,\n"
            f"  phone_number = EXCLUDED.phone_number,\n"
            f"  email = EXCLUDED.email,\n"
            f"  address_line_1 = EXCLUDED.address_line_1,\n"
            f"  address_line_2 = EXCLUDED.address_line_2,\n"
            f"  city = EXCLUDED.city,\n"
            f"  postcode = EXCLUDED.postcode;"
        )
        lines.append("")

        ref_num += 1
        referral_data.append((p, nhs, ref_num))

    if skipped:
        lines.append("-- SKIPPED (missing or invalid NHS number):")
        for name, nhs in skipped:
            lines.append(f"--   {name} (NHS: {nhs})")
        lines.append("")

    # ── mhs002_gp ──
    lines.extend([
        "-- =============================================================",
        "-- mhs002_gp  (GP Registration)",
        "-- =============================================================",
        "",
    ])

    for p, nhs, _ in referral_data:
        gp_onenote = p.get("_gp_onenote", {})
        gp_name = p.get("GP Practice Name") or gp_onenote.get("GP Practice") or None
        gp_email = p.get("GP Practice Email") or gp_onenote.get("GP Email") or None

        if gp_name:
            lines.append(
                f"INSERT INTO mhs002_gp "
                f"(nhs_number, gmp_code_reg, gp_practice_name, gp_contact_email)\n"
                f"SELECT {sql_val(nhs)}, 'V81999', {sql_val(gp_name)}, {sql_val(gp_email)}\n"
                f"WHERE NOT EXISTS (SELECT 1 FROM mhs002_gp WHERE nhs_number = {sql_val(nhs)});"
            )
            lines.append("")

    # ── mhs101_referral ──
    lines.extend([
        "-- =============================================================",
        "-- mhs101_referral  (Referral)",
        "-- =============================================================",
        "",
    ])

    for p, nhs, num in referral_data:
        service_request_id = f"REF-B2-{num:03d}"
        icb_name = p.get("ICB", "")
        icb_code = ICB_LOOKUP.get(icb_name)

        if not icb_code:
            logger.warning(f"  Unknown ICB '{icb_name}' for {p['Patient Name']}")
            icb_code = "QNX"  # Default to Sussex

        referral_date = parse_referral_date(p.get("Referral Date", "")) or "1900-01-01"

        # Add referrer info as SQL comment
        referrer_name = p.get("Referrer Name", "")
        referrer_email = p.get("Referrer Email", "")
        referrer_org = p.get("Referrer Organisation", "")
        if referrer_name:
            lines.append(
                f"-- Referrer: {referrer_name}"
                f"{', ' + referrer_email if referrer_email else ''}"
                f"{' (' + referrer_org + ')' if referrer_org else ''}"
            )

        lines.append(
            f"INSERT INTO mhs101_referral "
            f"(service_request_id, nhs_number, org_id_comm, "
            f"referral_request_received_date, source_of_referral_mh, "
            f"prim_reason_referral_mh, reason_oat, patient_status_code)\n"
            f"SELECT {sql_val(service_request_id)}, {sql_val(nhs)}, {sql_val(icb_code)}, "
            f"{sql_val(referral_date)}, 'A1', '24', '14', '01'\n"
            f"WHERE NOT EXISTS (SELECT 1 FROM mhs101_referral WHERE service_request_id = {sql_val(service_request_id)});"
        )
        lines.append("")

    # ── mhs102_service_type_referred_to ──
    lines.extend([
        "-- =============================================================",
        "-- mhs102_service_type_referred_to  (Service Type)",
        "-- =============================================================",
        "",
    ])

    for _, _, num in referral_data:
        service_request_id = f"REF-B2-{num:03d}"
        lines.append(
            f"INSERT INTO mhs102_service_type_referred_to "
            f"(service_request_id, serv_team_type_ref_to_mh)\n"
            f"SELECT {sql_val(service_request_id)}, 'C04'\n"
            f"WHERE NOT EXISTS (SELECT 1 FROM mhs102_service_type_referred_to "
            f"WHERE service_request_id = {sql_val(service_request_id)});"
        )
        lines.append("")

    lines.append("COMMIT;")
    lines.append("")

    with open(sql_path, "w") as f:
        f.write("\n".join(lines))

    logger.info(
        f"Phase E: Generated {sql_path} — "
        f"{len(referral_data)} patients, {len(skipped)} skipped"
    )
    for name, nhs in skipped:
        logger.warning(f"  SKIPPED: {name} (NHS: {nhs})")


# ── Write merged CSV ────────────────────────────────────────────────────────


CSV_COLUMNS = [
    "Patient Name",
    "Folder Name",
    "DOB",
    "NHS Number",
    "Patient Email",
    "Patient Phone Number",
    "Address Line 1",
    "Address Line 2",
    "City",
    "Post Code",
    "GP Practice Name",
    "GP Practice Email",
    "Referrer Organisation",
    "Referrer Name",
    "Referrer Email",
    "Referral Date",
    "ICB",
]


def write_merged_csv(patients, output_dir, suffix=""):
    """Write all_patients_batch2.csv with merged data."""
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"all_patients_batch2{suffix}.csv"

    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for p in patients:
            row = {col: p.get(col, "") for col in CSV_COLUMNS}
            writer.writerow(row)

    logger.info(f"Wrote {csv_path} ({len(patients)} rows)")


# ── Main ────────────────────────────────────────────────────────────────────


def main():
    if len(sys.argv) < 2:
        print(
            f"Usage: {sys.argv[0]} <batch2-root-dir>",
            file=sys.stderr,
        )
        sys.exit(1)

    root = Path(sys.argv[1])
    xlsx_path = root / "referral_register_batch2.xlsx"
    files_dir = root / "Patient Files"
    output_dir = Path("documents/batch2")

    if not xlsx_path.is_file():
        print(f"Error: Excel file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Phase A
    patients = load_excel(xlsx_path)

    # Phase B
    resolve_folders(patients, files_dir)

    # Phase C
    extract_onenote(patients)

    # Phase D
    cross_reference(patients, output_dir)

    # Phase E — full outputs (all patients)
    generate_sql(patients, output_dir)
    write_merged_csv(patients, output_dir)

    # Phase F — filtered outputs (no discrepancies)
    discrepancy_folders = get_discrepancy_folders(output_dir / "discrepancy_report.csv")
    clean_patients = [
        p for p in patients
        if p["Folder Name"] not in discrepancy_folders
        and p.get("ICB", "") and ICB_LOOKUP.get(p["ICB"])
    ]
    logger.info(
        f"Phase F: {len(clean_patients)} clean patients "
        f"({len(patients) - len(clean_patients)} with discrepancies excluded)"
    )
    generate_sql(clean_patients, output_dir, suffix="_clean")
    write_merged_csv(clean_patients, output_dir, suffix="_clean")

    logger.info("Done!")


if __name__ == "__main__":
    main()
